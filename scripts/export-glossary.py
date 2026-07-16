"""Export data/semantic-glossary.xlsx → JSON (+ web copy)."""

from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "semantic-glossary.xlsx"
JSON_OUT = ROOT / "data" / "semantic-glossary.json"
WEB_COPY = ROOT / "apps" / "web" / "src" / "data" / "semantic-glossary.json"


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    entries = []
    for r in rows[1:]:
        zh = str(r[0] or "").strip()
        en = str(r[1] or "").strip()
        note = str(r[2] or "").strip() if r[2] else ""
        if zh and en:
            entries.append({"zh": zh, "en": en, "note": note})

    payload = {
        "version": 1,
        "source": "semantic-glossary.xlsx",
        "count": len(entries),
        "entries": entries,
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    JSON_OUT.write_text(text, encoding="utf-8")
    WEB_COPY.parent.mkdir(parents=True, exist_ok=True)
    WEB_COPY.write_text(text, encoding="utf-8")
    print(f"Wrote {len(entries)} entries → {JSON_OUT}")
    print(f"Copied → {WEB_COPY}")


if __name__ == "__main__":
    main()
